from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import TableInfo, WarningInfo
from .access_reader import suggest_mysql_identifier

def _autosize_worksheet(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def _style_worksheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    _autosize_worksheet(ws)


def build_report_frames(tables: list[TableInfo], warnings: list[WarningInfo]) -> dict[str, pd.DataFrame]:
    total_rows = sum(t.row_count or 0 for t in tables)
    total_columns = sum(len(t.columns) for t in tables)

    warning_count = sum(1 for item in warnings if item.level == "warning")
    info_count = sum(1 for item in warnings if item.level == "info")

    pk_formal_count = sum(1 for table in tables if table.primary_key_source == "formal")
    pk_unique_index_count = sum(1 for table in tables if table.primary_key_source == "unique_index")
    pk_candidate_count = sum(1 for table in tables if table.primary_key_source == "candidate")
    pk_none_count = sum(1 for table in tables if table.primary_key_source == "none")

    summary_df = pd.DataFrame(
        [
            {"Metric": "Tables", "Value": len(tables)},
            {"Metric": "Columns", "Value": total_columns},
            {"Metric": "Rows", "Value": total_rows},
            {"Metric": "Warnings", "Value": warning_count},
            {"Metric": "Info", "Value": info_count},
            {"Metric": "Total notes", "Value": len(warnings)},
            {"Metric": "PK formal", "Value": pk_formal_count},
            {"Metric": "PK unique_index", "Value": pk_unique_index_count},
            {"Metric": "PK candidate", "Value": pk_candidate_count},
            {"Metric": "PK none", "Value": pk_none_count},
        ]
    )

    tables_df = pd.DataFrame(
        [
            {
                "Table": table.table_name,
                "Recommended MySQL Table": suggest_mysql_identifier(table.table_name),
                "Rows": table.row_count,
                "Columns": len(table.columns),
                "Primary Key Status": table.primary_key_source,
                "Primary Key Columns": ", ".join(table.primary_keys),
            }
            for table in tables
        ]
    )

    primary_keys_df = pd.DataFrame(
        [
            {
                "Table": table.table_name,
                "PK Status": table.primary_key_source,
                "PK Columns": ", ".join(table.primary_keys),
                "Rows": table.row_count,
                "Columns": len(table.columns),
                "Note": (
                    "Formal primary key detected."
                    if table.primary_key_source == "formal"
                    else "Unique index detected; may represent primary key."
                    if table.primary_key_source == "unique_index"
                    else "Possible primary key candidate detected."
                    if table.primary_key_source == "candidate"
                    else "No primary key or obvious candidate detected."
                ),
            }
            for table in tables
        ]
    )

    cleanup_rows = []

    for table in tables:
        reasons = []

        table_name_lower = table.table_name.lower()

        if table.row_count == 0:
            reasons.append("Empty table")

        if table.primary_key_source == "none":
            reasons.append("No primary key detected")

        if "importerrors" in table_name_lower or "import errors" in table_name_lower:
            reasons.append("Access import error table")

        if "copy of" in table_name_lower or "copy2 of" in table_name_lower or "kopija" in table_name_lower:
            reasons.append("Backup/copy table")

        if "backup" in table_name_lower or "_bak" in table_name_lower or " bak" in table_name_lower:
            reasons.append("Backup table")

        if "temp" in table_name_lower or "tmp" in table_name_lower:
            reasons.append("Temporary table")

        if "test" in table_name_lower:
            reasons.append("Test table")

        if "staro" in table_name_lower or "old" in table_name_lower:
            reasons.append("Old/legacy copy table")

        if reasons:
            high_risk_keywords = [
                "Access import error table",
                "Backup/copy table",
                "Backup table",
                "Temporary table",
                "Test table",
                "Old/legacy copy table",
            ]

            if any(reason in high_risk_keywords for reason in reasons):
                cleanup_priority = "High"
                suggested_action = "Review before migration. Exclude if this is not a production table."
            elif "No primary key detected" in reasons:
                cleanup_priority = "Medium"
                suggested_action = "Review table structure before migration. Consider adding or confirming a primary key."
            elif reasons == ["Empty table"]:
                cleanup_priority = "Low"
                suggested_action = "Table is empty. Keep it if it is a valid application/domain table."
            else:
                cleanup_priority = "Low"
                suggested_action = "Review before migration."

            cleanup_rows.append(
                {
                    "Table": table.table_name,
                    "Rows": table.row_count,
                    "Columns": len(table.columns),
                    "PK Status": table.primary_key_source,
                    "Cleanup Priority": cleanup_priority,
                    "Reasons": "; ".join(reasons),
                    "Suggested Action": suggested_action,
                }
            )

    cleanup_candidates_df = pd.DataFrame(cleanup_rows)

    if cleanup_candidates_df.empty:
        cleanup_candidates_df = pd.DataFrame(
            [
                {
                    "Table": None,
                    "Rows": None,
                    "Columns": None,
                    "PK Status": None,
                    "Cleanup Priority": None,
                    "Reasons": "No cleanup candidates detected.",
                    "Suggested Action": None,
                }
            ]
        )

    migration_plan_rows = []

    for table in tables:
        table_name_lower = table.table_name.lower()
        reasons = []

        is_import_error = "importerrors" in table_name_lower or "import errors" in table_name_lower
        is_copy_table = (
            "copy of" in table_name_lower
            or "copy2 of" in table_name_lower
            or "kopija" in table_name_lower
        )
        is_backup_table = "backup" in table_name_lower or "_bak" in table_name_lower or " bak" in table_name_lower
        is_temp_table = "temp" in table_name_lower or "tmp" in table_name_lower
        is_test_table = "test" in table_name_lower
        is_old_table = "staro" in table_name_lower or "old" in table_name_lower
        is_empty = table.row_count == 0
        has_no_pk = table.primary_key_source == "none"

        if is_import_error:
            reasons.append("Access import error table")
        if is_copy_table:
            reasons.append("Backup/copy table")
        if is_backup_table:
            reasons.append("Backup table")
        if is_temp_table:
            reasons.append("Temporary table")
        if is_test_table:
            reasons.append("Test table")
        if is_old_table:
            reasons.append("Old/legacy copy table")
        if is_empty:
            reasons.append("Empty table")
        if has_no_pk:
            reasons.append("No primary key detected")

        if is_import_error or is_copy_table or is_backup_table or is_temp_table or is_test_table or is_old_table:
            recommendation = "Exclude by default"
            action = "Exclude from migration unless confirmed as a real production table."
        elif is_empty:
            recommendation = "Review"
            action = "Review with application owner. Keep if this is a valid empty domain/application table."
        elif has_no_pk:
            recommendation = "Review before migration"
            action = "Review table structure and consider adding or confirming a primary key."
        else:
            recommendation = "Migrate"
            action = "Table looks suitable for migration based on current checks."

        migration_plan_rows.append(
            {
                "Table": table.table_name,
                "Recommended MySQL Table": suggest_mysql_identifier(table.table_name),
                "Recommendation": recommendation,
                "Rows": table.row_count,
                "Columns": len(table.columns),
                "PK Status": table.primary_key_source,
                "PK Columns": ", ".join(table.primary_keys),
                "Reasons": "; ".join(reasons) if reasons else "No major issue detected",
                "Suggested Action": action,
            }
        )

    migration_plan_df = pd.DataFrame(migration_plan_rows)

    columns_df = pd.DataFrame(
        [
            {
                "Table": column.table_name,
                "Recommended MySQL Table": suggest_mysql_identifier(column.table_name),
                "Column": column.column_name,
                "Recommended MySQL Column": suggest_mysql_identifier(column.column_name),
                "Ordinal": column.ordinal_position,
                "Access/ODBC Type": column.type_name,
                "ODBC Data Type": column.data_type,
                "Size": column.column_size,
                "Decimal Digits": column.decimal_digits,
                "Nullable": column.nullable,
                "Empty Values": column.empty_count,
                "Filled Values": column.filled_count,
                "Fill Rate %": column.fill_rate_percent,
                "Suggested MySQL Type": column.mysql_type,
            }
            for table in tables
            for column in table.columns
        ]
    )

    type_mapping_df = columns_df[["Access/ODBC Type", "Suggested MySQL Type"]].drop_duplicates() if not columns_df.empty else pd.DataFrame(columns=["Access/ODBC Type", "Suggested MySQL Type"])

    warnings_df = pd.DataFrame(
        [
            {
                "Level": warning.level,
                "Table": warning.table_name,
                "Column": warning.column_name,
                "Message": warning.message,
            }
            for warning in warnings
        ]
    )
    if warnings_df.empty:
        warnings_df = pd.DataFrame([{"Level": "info", "Table": None, "Column": None, "Message": "No warnings generated in this starter scan."}])

    return {
        "Summary": summary_df,
        "Migration Plan": migration_plan_df,
        "Tables": tables_df,
        "Primary Keys": primary_keys_df,
        "Cleanup Candidates": cleanup_candidates_df,
        "Columns": columns_df,
        "Type Mapping": type_mapping_df,
        "Warnings": warnings_df,
    }


def write_excel_report(tables: list[TableInfo], warnings: list[WarningInfo], output_path: str | Path) -> Path:
    output = Path(output_path)
    frames = build_report_frames(tables, warnings)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in frames.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        for ws in workbook.worksheets:
            _style_worksheet(ws)

    return output
